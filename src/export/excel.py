"""Export the comparison DB to a 3-sheet Excel workbook.

Sheet 1 — Price Comparison: all normalized prices, one row per (vendor, line)
Sheet 2 — Vendor Terms: freight and discount data
Sheet 3 — Questionnaire: qualification answers
"""

import io
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.db.connection import get_db

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center")

_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")  # low confidence
_BAD_FILL  = PatternFill("solid", fgColor="FFCCCC")  # very low / missing


def _style_header(ws, row: int, cols: list[str]) -> None:
    for col_idx, val in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


async def build_excel(rfx_id: str = "RFX-001") -> bytes:
    """Build the 3-sheet Excel workbook and return raw bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Price Comparison ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Price Comparison"

    headers1 = ["Vendor", "Line ID", "Description", "Price INR", "Unit (Raw)",
                 "Currency", "Confidence", "Flags", "Extraction Status"]
    _style_header(ws1, 1, headers1)

    async with get_db() as db:
        cursor = await db.execute(
            """
            SELECT vendor_id, line_id, description, price_inr, unit_raw,
                   currency_raw, confidence, flags, extraction_status
            FROM comparison WHERE rfx_id = ?
            ORDER BY line_id, vendor_id
            """,
            (rfx_id,),
        )
        rows = await cursor.fetchall()

    for r in rows:
        flags = json.loads(r["flags"]) if r["flags"] else []
        conf = r["confidence"] or 0.0
        row_data = [
            r["vendor_id"],
            r["line_id"],
            r["description"],
            round(r["price_inr"], 2) if r["price_inr"] is not None else None,
            r["unit_raw"],
            r["currency_raw"],
            round(conf, 2),
            ", ".join(flags),
            r["extraction_status"],
        ]
        ws1.append(row_data)
        row_idx = ws1.max_row
        fill = _BAD_FILL if conf < 0.3 else (_WARN_FILL if conf < 0.7 else None)
        if fill:
            for col_idx in range(1, len(headers1) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    _auto_width(ws1)

    # ── Sheet 2: Vendor Terms ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Vendor Terms")
    headers2 = ["Vendor", "Freight (INR)", "Freight Notes", "Freight Unquantified",
                 "Discount Condition", "Discount %"]
    _style_header(ws2, 1, headers2)

    async with get_db() as db:
        cursor = await db.execute(
            """
            SELECT vendor_id, freight_inr, freight_notes, freight_unquantified,
                   discount_condition, discount_pct
            FROM vendor_terms WHERE rfx_id = ?
            ORDER BY vendor_id
            """,
            (rfx_id,),
        )
        rows = await cursor.fetchall()

    for r in rows:
        ws2.append([
            r["vendor_id"],
            r["freight_inr"],
            r["freight_notes"],
            "Yes" if r["freight_unquantified"] else "No",
            r["discount_condition"],
            r["discount_pct"],
        ])

    _auto_width(ws2)

    # ── Sheet 3: Questionnaire ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Questionnaire")
    headers3 = ["Vendor", "ISO Certified", "Rejection Rate (%)", "Lead Time (days)",
                 "Manufacturing Location", "Deviations", "Quote Validity (days)"]
    _style_header(ws3, 1, headers3)

    async with get_db() as db:
        cursor = await db.execute(
            """
            SELECT vendor_id, iso_certified, rejection_rate, lead_time_days,
                   manufacturing_location, deviations, quote_validity_days
            FROM questionnaire WHERE rfx_id = ?
            ORDER BY vendor_id
            """,
            (rfx_id,),
        )
        rows = await cursor.fetchall()

    for r in rows:
        ws3.append([
            r["vendor_id"],
            r["iso_certified"],
            r["rejection_rate"],
            r["lead_time_days"],
            r["manufacturing_location"],
            r["deviations"],
            r["quote_validity_days"],
        ])

    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
